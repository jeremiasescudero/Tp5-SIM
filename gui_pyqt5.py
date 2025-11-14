"""
Aplicación PyQt5 para simulación de biblioteca con tabla de eventos
CON INTERFAZ DE CONFIGURACIÓN MÍNIMA Y PARAMETRIZADA.
Incluye:
1. Parametrización de los 4 datos requeridos (llegadas, objetivos, consulta, retiro).
2. Uso de copia profunda para corregir el bug de los estados de los clientes en la tabla.
3. Eliminación de la columna "Estado" y actualización de índices.
4. Resaltado en rojo del próximo evento.
5. CORRECCIÓN DEL ERROR: NameError: name 'clientes_dict' is not defined.
6. Parametrización de la cantidad de iteraciones (Máx: 100000).
7. Exportación de las Tablas del Método de Euler a una hoja separada en Excel.
8. Límite de 300 filas visibles en la interfaz, exportando todas a Excel.
9. CORRECCIÓN: Priorización de la simulación por ITERACIONES (max_iteraciones) sobre el tiempo (tiempo_maximo).
10. AÑADIDO: Pestaña de Resultados/Métricas en la interfaz principal.
"""
import sys
import random
import heapq
import math
from enum import Enum
from dataclasses import dataclass
from typing import List, Optional, Dict
import copy 

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QTableWidget, QTableWidgetItem, QHeaderView,
    QLabel, QSpinBox, QDoubleSpinBox, QGroupBox, QFormLayout,
    QProgressBar, QMessageBox, QGridLayout, QScrollArea, QFileDialog,
    QDialog, QLineEdit, QTabWidget
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal
from PyQt5.QtGui import QColor, QFont

try:
    import openpyxl
    from openpyxl.styles import Font as ExcelFont, PatternFill, Alignment
    OPENPYXL_DISPONIBLE = True
except ImportError:
    OPENPYXL_DISPONIBLE = False


# ==================== INTEGRADOR DE EULER ====================

class IntegradorEuler:
    """INTEGRACIÓN NUMÉRICA POR MÉTODO DE EULER"""
    def __init__(self, h: float, K: int, p_inicial: float = 0):
        self.h = h
        self.K = K
        self.p = p_inicial
        self.t = 0.0
        self.tabla_euler = [{'t': 0.0, 'p': p_inicial, 'dp/dt': self.derivada(p_inicial, 0.0)}] 

    def derivada(self, p: float, t: float) -> float:
        return self.K / 5.0
    
    def paso(self) -> float:
        derivada_actual = self.derivada(self.p, self.t)
        self.p = self.p + self.h * derivada_actual
        self.t += self.h
        self.tabla_euler.append({'t': self.t, 'p': self.p, 'dp/dt': self.derivada(self.p, self.t)}) 
        return self.p
    
    def integrar_hasta_paginas(self, paginas_objetivo: float) -> float:
        max_pasos = int(paginas_objetivo / self.h * 1.5)
        pasos_dados = 0
        while self.p < paginas_objetivo and pasos_dados < max_pasos: 
            self.paso()
            pasos_dados += 1
        return self.t


# ==================== ENUMS Y DATACLASSES ====================

class TipoEvento(Enum):
    INICIALIZACION = "inicializacion"
    LLEGADA_CLIENTE = "llegada_cliente"
    FIN_ATENCION = "fin_atencion_cliente"
    FIN_LECTURA = "fin_lectura"
    FIN_SIMULACION = "fin_simulacion"
class TipoObjetivo(Enum):
    PEDIR_LIBRO = "Pedir libro"
    DEVOLVER_LIBRO = "Devolver libro"
    CONSULTAR = "Consultar"
class EstadoEmpleado(Enum):
    LIBRE = "Libre"
    OCUPADO = "Ocupado"
@dataclass
class Evento:
    tiempo: float
    tipo: TipoEvento
    datos: dict
    def __lt__(self, otro):
        return self.tiempo < otro.tiempo
@dataclass
class Cliente:
    id: int
    hora_llegada: float
    objetivo: TipoObjetivo
    rnd_objetivo: float
    rnd_tiempo_consulta: float = 0.0
    tiempo_consulta: float = 0.0
    rnd_tiempo_busqueda: float = 0.0
    tiempo_busqueda: float = 0.0
    rnd_tiempo_devolucion: float = 0.0
    tiempo_devolucion: float = 0.0
    fin_atencion_emp1: Optional[float] = None
    fin_atencion_emp2: Optional[float] = None
    se_retira: bool = False
    rnd_decision: Optional[float] = None
    paginas_a_leer: int = 0
    rnd_paginas: Optional[float] = None
    tiempo_lectura: float = 0.0
    fin_lectura: Optional[float] = None
    estado: str = "En cola"
    hora_salida: Optional[float] = None
    tabla_euler: Optional[List[Dict[str, float]]] = None
class Empleado:
    def __init__(self, id: int):
        self.id = id
        self.estado = EstadoEmpleado.LIBRE
        self.cliente_actual: Optional[Cliente] = None
        self.hora_fin_atencion: Optional[float] = None
        self.tiempo_acumulado_atencion = 0.0
        self.tiempo_acumulado_ocioso = 0.0
        self.ultimo_cambio_estado = 0.0
    def esta_libre(self):
        return self.estado == EstadoEmpleado.LIBRE
    def atender(self, cliente: Cliente, tiempo_fin: float, reloj: float):
        if self.estado == EstadoEmpleado.LIBRE:
            self.tiempo_acumulado_ocioso += reloj - self.ultimo_cambio_estado
        self.estado = EstadoEmpleado.OCUPADO
        self.cliente_actual = cliente
        self.hora_fin_atencion = tiempo_fin
        self.ultimo_cambio_estado = reloj
    def liberar(self, reloj: float):
        if self.estado == EstadoEmpleado.OCUPADO:
            self.tiempo_acumulado_atencion += reloj - self.ultimo_cambio_estado
        self.estado = EstadoEmpleado.LIBRE
        cliente = self.cliente_actual
        self.cliente_actual = None
        self.hora_fin_atencion = None
        self.ultimo_cambio_estado = reloj
        return cliente


# ==================== MOTOR DE SIMULACIÓN ====================

class Simulacion:
    """Motor de simulación de eventos discretos"""

    def __init__(self, parametros: dict): 
        
        # --- PARÁMETROS CONFIGURABLES ---
        self.tiempo_entre_llegadas = parametros.get('tiempo_entre_llegadas', 4.0)
        self.prob_pedir_libro = parametros.get('prob_pedir_libro', 0.45)
        self.prob_devolver_libro = parametros.get('prob_devolver_libro', 0.45)
        self.prob_consultar = parametros.get('prob_consultar', 0.10)
        self.tiempo_consulta_min = parametros.get('tiempo_consulta_min', 2.0)
        self.tiempo_consulta_max = parametros.get('tiempo_consulta_max', 5.0)
        self.prob_retirarse = parametros.get('prob_retirarse', 0.60)
        self.max_iteraciones = parametros.get('max_iteraciones', 10000) 
        
        # --- PARÁMETROS DE K PARA INTEGRACIÓN ---
        self.K_100_200 = parametros.get('K_100_200', 100)
        self.K_200_300 = parametros.get('K_200_300', 90)
        self.K_300_plus = parametros.get('K_300_plus', 70)

        # --- PARÁMETROS FIJOS (VALORES PREDETERMINADOS) ---
        self.media_busqueda = 6.0
        self.tiempo_devolucion_min = 1.5
        self.tiempo_devolucion_max = 2.5
        self.paginas_min = 100
        self.paginas_max = 350
        self.capacidad_maxima = 20
        self.tiempo_maximo = 480.0
        self.h_euler = 0.1

        # Estado de la simulación
        self.reloj = 0.0
        self.eventos = []
        self.clientes_activos: List[Cliente] = []
        self.cola_espera: List[Cliente] = []
        self.clientes_leyendo: List[Cliente] = []
        self.empleados = [Empleado(1), Empleado(2)]
        self.contador_clientes = 0
        self.numero_fila = 0
        self.ultimo_reloj = 0.0

        # Acumuladores
        self.ac_tiempo_permanencia = 0.0
        self.total_clientes_atendidos = 0
        self.total_clientes_leyendo = 0
        self.total_clientes_generados = 0
        self.total_clientes_rechazados = 0
        self.biblioteca_cerrada = False
        self.tiempo_biblioteca_cerrada_ac = 0.0
        self.tiempo_inicio_cerrada: Optional[float] = None

        self.historial_filas: List[dict] = []
        self.tablas_euler_clientes: Dict[int, List[Dict[str, float]]] = {} 

    def determinar_K(self, num_paginas: int) -> int:
        if 100 <= num_paginas <= 200:
            return self.K_100_200
        elif 200 < num_paginas <= 300:
            return self.K_200_300
        else:
            return self.K_300_plus

    def generar_objetivo(self) -> tuple:
        rnd = random.random()
        if rnd < self.prob_pedir_libro:
            return TipoObjetivo.PEDIR_LIBRO, rnd
        elif rnd < self.prob_pedir_libro + self.prob_devolver_libro:
            return TipoObjetivo.DEVOLVER_LIBRO, rnd
        else:
            return TipoObjetivo.CONSULTAR, rnd

    def generar_tiempo_consulta(self) -> tuple:
        rnd = random.random()
        tiempo = self.tiempo_consulta_min + (self.tiempo_consulta_max - self.tiempo_consulta_min) * rnd
        return tiempo, rnd

    def generar_tiempo_busqueda(self) -> tuple:
        rnd = random.random()
        media = self.media_busqueda if self.media_busqueda > 0 else 1e-6
        tiempo = -media * math.log(1 - rnd)
        return tiempo, rnd

    def generar_tiempo_devolucion(self) -> tuple:
        rnd = random.random()
        tiempo = self.tiempo_devolucion_min + (self.tiempo_devolucion_max - self.tiempo_devolucion_min) * rnd
        return tiempo, rnd

    def agregar_evento(self, evento: Evento):
        heapq.heappush(self.eventos, evento)

    def proximo_evento(self) -> Optional[Evento]:
        if self.eventos:
            return heapq.heappop(self.eventos)
        return None

    def actualizar_tiempo_cerrada(self):
        if self.biblioteca_cerrada and self.tiempo_inicio_cerrada is not None:
            self.tiempo_biblioteca_cerrada_ac += self.reloj - self.tiempo_inicio_cerrada
            self.tiempo_inicio_cerrada = self.reloj

    def iniciar(self):
        self.numero_fila = 0
        self.historial_filas.append(self._capturar_estado(
            Evento(tiempo=0.0, tipo=TipoEvento.INICIALIZACION, datos={})
        ))
        self.numero_fila += 1
        self.ultimo_reloj = 0.0

        tiempo_primera_llegada = self.tiempo_entre_llegadas
        self.agregar_evento(Evento(
            tiempo=tiempo_primera_llegada,
            tipo=TipoEvento.LLEGADA_CLIENTE,
            datos={}
        ))
        
        # Se elimina el evento FIN_SIMULACION

    def procesar_llegada_cliente(self, evento: Evento):
        self.total_clientes_generados += 1
        
        if len(self.clientes_activos) >= self.capacidad_maxima:
            self.total_clientes_rechazados += 1
            
            proxima_llegada = self.reloj + self.tiempo_entre_llegadas
            
            self.agregar_evento(Evento(
                tiempo=proxima_llegada,
                tipo=TipoEvento.LLEGADA_CLIENTE,
                datos={}
            ))
            
            cliente_rechazado = Cliente(
                id=self.total_clientes_generados, 
                hora_llegada=self.reloj,
                objetivo=TipoObjetivo.CONSULTAR, 
                rnd_objetivo=0.0,
                estado="RECHAZADO" 
            )
            evento.datos['cliente'] = cliente_rechazado
            return

        self.contador_clientes += 1
        objetivo, rnd_objetivo = self.generar_objetivo()
        cliente = Cliente(
            id=self.contador_clientes,
            hora_llegada=self.reloj,
            objetivo=objetivo,
            rnd_objetivo=rnd_objetivo
        )

        self.clientes_activos.append(cliente)
        evento.datos['cliente'] = cliente
        
        if len(self.clientes_activos) >= self.capacidad_maxima:
            self.biblioteca_cerrada = True
            self.tiempo_inicio_cerrada = self.reloj

        empleado_libre = self._obtener_empleado_libre()
        if empleado_libre:
            self._atender_cliente(cliente, empleado_libre)
        else:
            self.cola_espera.append(cliente)
            cliente.estado = "En cola"

        proxima_llegada = self.reloj + self.tiempo_entre_llegadas
        
        self.agregar_evento(Evento(
            tiempo=proxima_llegada,
            tipo=TipoEvento.LLEGADA_CLIENTE,
            datos={}
        ))

    def _obtener_empleado_libre(self) -> Optional[Empleado]:
        for emp in self.empleados:
            if emp.esta_libre():
                return emp
        return None

    def _atender_cliente(self, cliente: Cliente, empleado: Empleado):
        cliente.estado = "Siendo atendido"

        if cliente.objetivo == TipoObjetivo.CONSULTAR:
            tiempo_atencion, rnd = self.generar_tiempo_consulta()
            cliente.tiempo_consulta = tiempo_atencion
            cliente.rnd_tiempo_consulta = rnd

        elif cliente.objetivo == TipoObjetivo.PEDIR_LIBRO:
            tiempo_atencion, rnd = self.generar_tiempo_busqueda()
            cliente.tiempo_busqueda = tiempo_atencion
            cliente.rnd_tiempo_busqueda = rnd

        else:
            tiempo_atencion, rnd = self.generar_tiempo_devolucion()
            cliente.tiempo_devolucion = tiempo_atencion
            cliente.rnd_tiempo_devolucion = rnd

        tiempo_fin = self.reloj + tiempo_atencion

        if empleado.id == 1:
            cliente.fin_atencion_emp1 = tiempo_fin
        else:
            cliente.fin_atencion_emp2 = tiempo_fin

        empleado.atender(cliente, tiempo_fin, self.reloj)

        self.agregar_evento(Evento(
            tiempo=tiempo_fin,
            tipo=TipoEvento.FIN_ATENCION,
            datos={'cliente': cliente, 'empleado': empleado}
        ))

    def procesar_fin_atencion(self, evento: Evento):
        cliente = evento.datos['cliente']
        empleado = evento.datos['empleado']

        empleado.liberar(self.reloj)

        if cliente.objetivo == TipoObjetivo.PEDIR_LIBRO:
            rnd_decision = random.random()
            cliente.rnd_decision = rnd_decision

            if rnd_decision < self.prob_retirarse:
                cliente.se_retira = True
                cliente.estado = "Fuera del sistema" 
                cliente.hora_salida = self.reloj
                self._cliente_sale(cliente)
            else:
                cliente.se_retira = False
                cliente.estado = "Leyendo" 

                rnd_paginas = random.random()
                cliente.rnd_paginas = rnd_paginas
                cliente.paginas_a_leer = int(self.paginas_min +
                                                 (self.paginas_max - self.paginas_min) * rnd_paginas)

                K = self.determinar_K(cliente.paginas_a_leer)
                integrador = IntegradorEuler(h=self.h_euler, K=K, p_inicial=0)
                cliente.tiempo_lectura = integrador.integrar_hasta_paginas(cliente.paginas_a_leer)
                cliente.fin_lectura = self.reloj + cliente.tiempo_lectura
                
                # Guardar la tabla de Euler
                cliente.tabla_euler = integrador.tabla_euler
                self.tablas_euler_clientes[cliente.id] = integrador.tabla_euler


                self.clientes_leyendo.append(cliente)
                self.total_clientes_leyendo += 1

                self.agregar_evento(Evento(
                    tiempo=cliente.fin_lectura,
                    tipo=TipoEvento.FIN_LECTURA,
                    datos={'cliente': cliente}
                ))
        else:
            cliente.estado = "Fuera del sistema" 
            cliente.hora_salida = self.reloj
            self._cliente_sale(cliente)

        if self.cola_espera and empleado.esta_libre():
            siguiente = self.cola_espera.pop(0)
            self._atender_cliente(siguiente, empleado)

    def procesar_fin_lectura(self, evento: Evento):
        cliente = evento.datos['cliente']

        if cliente in self.clientes_leyendo:
            self.clientes_leyendo.remove(cliente)

        # El cliente ahora debe DEVOLVER el libro
        # Cambiar su objetivo a devolver libro
        cliente.objetivo = TipoObjetivo.DEVOLVER_LIBRO

        # Buscar empleado libre
        empleado_libre = self._obtener_empleado_libre()
        if empleado_libre:
            # Atender inmediatamente para devolución
            cliente.estado = "Siendo atendido"
            self._atender_cliente(cliente, empleado_libre)
        else:
            # Si no hay empleado libre, entrar a la cola
            cliente.estado = "En cola"
            self.cola_espera.append(cliente)

    def _cliente_sale(self, cliente: Cliente):
        if cliente in self.clientes_activos:
            self.clientes_activos.remove(cliente)

        if cliente.hora_salida:
            tiempo_permanencia = cliente.hora_salida - cliente.hora_llegada
            self.ac_tiempo_permanencia += tiempo_permanencia

        self.total_clientes_atendidos += 1

        if self.biblioteca_cerrada and len(self.clientes_activos) < self.capacidad_maxima:
            self.biblioteca_cerrada = False
            self.tiempo_inicio_cerrada = None

    def ejecutar_paso(self) -> Optional[dict]:
        
        # Detener la simulación si se supera el máximo de iteraciones
        if self.numero_fila > self.max_iteraciones:
            # Procesar el último estado antes de terminar forzadamente
            if self.biblioteca_cerrada and self.tiempo_inicio_cerrada is not None:
                self.tiempo_biblioteca_cerrada_ac += self.tiempo_maximo - self.reloj 
            
            fila_datos = self._capturar_estado(Evento(tiempo=self.reloj, tipo=TipoEvento.FIN_SIMULACION, datos={}))
            self.historial_filas.append(fila_datos)
            return None
            
        evento = self.proximo_evento()
        
        # El bucle termina si no hay más eventos en la cola
        if not evento:
            return None 

        # Si el reloj excede el tiempo máximo, actualizamos el tiempo de cerrada para las métricas, 
        # pero permitimos que los eventos de servicio/lectura pendientes continúen si es necesario.
        if self.reloj > self.tiempo_maximo and self.tiempo_inicio_cerrada is not None:
            # Esto se mantiene solo para registrar el tiempo que la biblioteca estuvo "cerrada por capacidad"
            tiempo_transcurrido = evento.tiempo - self.reloj
            self.tiempo_biblioteca_cerrada_ac += tiempo_transcurrido

        self.reloj = evento.tiempo
        self.ultimo_reloj = self.reloj
        
        if evento.tipo == TipoEvento.LLEGADA_CLIENTE:
            self.procesar_llegada_cliente(evento)
        elif evento.tipo == TipoEvento.FIN_ATENCION:
            self.procesar_fin_atencion(evento)
        elif evento.tipo == TipoEvento.FIN_LECTURA:
            self.procesar_fin_lectura(evento)

        fila_datos = self._capturar_estado(evento)

        self.historial_filas.append(fila_datos)
        self.numero_fila += 1 

        return fila_datos

    def ejecutar_completa(self):
        self.iniciar()
        while True:
            resultado = self.ejecutar_paso()
            if resultado is None:
                break
        return self.historial_filas

    def _capturar_estado(self, evento: Evento) -> dict:
        proximos = sorted(self.eventos, key=lambda e: e.tiempo)[:3]
        cliente_actual = evento.datos.get('cliente')

        evento_str = evento.tipo.value
        if cliente_actual and evento.tipo != TipoEvento.INICIALIZACION:
            evento_str = f"{evento.tipo.value} C{cliente_actual.id}"

        rnd_busqueda = cliente_actual.rnd_tiempo_busqueda if cliente_actual and cliente_actual.objetivo == TipoObjetivo.PEDIR_LIBRO and cliente_actual.rnd_tiempo_busqueda > 0 else ''
        tiempo_busqueda = cliente_actual.tiempo_busqueda if cliente_actual and cliente_actual.objetivo == TipoObjetivo.PEDIR_LIBRO and cliente_actual.tiempo_busqueda > 0 else ''
        rnd_devolucion = cliente_actual.rnd_tiempo_devolucion if cliente_actual and cliente_actual.objetivo == TipoObjetivo.DEVOLVER_LIBRO and cliente_actual.rnd_tiempo_devolucion > 0 else ''
        tiempo_devolucion = cliente_actual.tiempo_devolucion if cliente_actual and cliente_actual.objetivo == TipoObjetivo.DEVOLVER_LIBRO and cliente_actual.rnd_tiempo_devolucion > 0 else ''
        rnd_consulta = cliente_actual.rnd_tiempo_consulta if cliente_actual and cliente_actual.objetivo == TipoObjetivo.CONSULTAR and cliente_actual.rnd_tiempo_consulta > 0 else ''
        tiempo_consulta = cliente_actual.tiempo_consulta if cliente_actual and cliente_actual.objetivo == TipoObjetivo.CONSULTAR and cliente_actual.tiempo_consulta > 0 else ''

        proxima_llegada = next((e.tiempo for e in proximos if e.tipo == TipoEvento.LLEGADA_CLIENTE), '')

        objetivo_val = cliente_actual.objetivo.value if cliente_actual and cliente_actual.estado != "RECHAZADO" else ('RECHAZADO' if cliente_actual and cliente_actual.estado == "RECHAZADO" else '')
        rnd_obj_val = cliente_actual.rnd_objetivo if cliente_actual and cliente_actual.estado != "RECHAZADO" else ''

        clientes_copiados = [copy.deepcopy(c) for c in self.clientes_activos]

        # NUEVO: Capturar TODOS los tiempos de fin de atención pendientes (de TODOS los empleados)
        fin_atencion_alq1 = ''
        fin_atencion_alq2 = ''
        fin_atencion_dev1 = ''
        fin_atencion_dev2 = ''
        fin_atencion_cons = ''

        # Buscar en empleados ocupados sus tiempos de finalización
        if self.empleados[0].hora_fin_atencion is not None and self.empleados[0].hora_fin_atencion > self.reloj:
            cliente_emp1 = self.empleados[0].cliente_actual
            if cliente_emp1:
                if cliente_emp1.objetivo == TipoObjetivo.PEDIR_LIBRO:
                    fin_atencion_alq1 = self.empleados[0].hora_fin_atencion
                elif cliente_emp1.objetivo == TipoObjetivo.DEVOLVER_LIBRO:
                    fin_atencion_dev1 = self.empleados[0].hora_fin_atencion
                elif cliente_emp1.objetivo == TipoObjetivo.CONSULTAR:
                    fin_atencion_cons = self.empleados[0].hora_fin_atencion

        if self.empleados[1].hora_fin_atencion is not None and self.empleados[1].hora_fin_atencion > self.reloj:
            cliente_emp2 = self.empleados[1].cliente_actual
            if cliente_emp2:
                if cliente_emp2.objetivo == TipoObjetivo.PEDIR_LIBRO:
                    fin_atencion_alq2 = self.empleados[1].hora_fin_atencion
                elif cliente_emp2.objetivo == TipoObjetivo.DEVOLVER_LIBRO:
                    fin_atencion_dev2 = self.empleados[1].hora_fin_atencion
                elif cliente_emp2.objetivo == TipoObjetivo.CONSULTAR and not fin_atencion_cons:
                    fin_atencion_cons = self.empleados[1].hora_fin_atencion

        return {
            'n': self.numero_fila,
            'evento': evento_str,
            'reloj': self.reloj,
            'tiempo_entre_llegadas': self.tiempo_entre_llegadas if evento.tipo == TipoEvento.LLEGADA_CLIENTE or evento.tipo == TipoEvento.INICIALIZACION else '',
            'proxima_llegada': proxima_llegada if evento.tipo != TipoEvento.INICIALIZACION else self.tiempo_entre_llegadas,
            'rnd_llegada': '',
            'rnd_objetivo': rnd_obj_val,
            'objetivo': objetivo_val,
            'rnd_busqueda': rnd_busqueda, 'tiempo_busqueda': tiempo_busqueda,
            'fin_atencion_alq1': fin_atencion_alq1,
            'fin_atencion_alq2': fin_atencion_alq2,
            'rnd_decision': cliente_actual.rnd_decision if cliente_actual and cliente_actual.rnd_decision is not None else '',
            'se_retira': 'Sí' if cliente_actual and cliente_actual.se_retira else ('No' if cliente_actual and cliente_actual.rnd_decision is not None else ''),
            'rnd_paginas': cliente_actual.rnd_paginas if cliente_actual and cliente_actual.rnd_paginas else '',
            'paginas': cliente_actual.paginas_a_leer if cliente_actual and cliente_actual.paginas_a_leer > 0 else '',
            'tiempo_lectura': cliente_actual.tiempo_lectura if cliente_actual and cliente_actual.tiempo_lectura > 0 else '',
            'rnd_devolucion': rnd_devolucion, 'tiempo_devolucion': tiempo_devolucion,
            'fin_atencion_dev1': fin_atencion_dev1,
            'fin_atencion_dev2': fin_atencion_dev2,
            'rnd_consulta': rnd_consulta, 'tiempo_consulta': tiempo_consulta,
            'fin_atencion_cons': fin_atencion_cons,
            'empleado1_estado': self.empleados[0].estado.value,
            'empleado1_ac_atencion': self.empleados[0].tiempo_acumulado_atencion,
            'empleado1_ac_ocioso': self.empleados[0].tiempo_acumulado_ocioso,
            'empleado2_estado': self.empleados[1].estado.value,
            'empleado2_ac_atencion': self.empleados[1].tiempo_acumulado_atencion,
            'empleado2_ac_ocioso': self.empleados[1].tiempo_acumulado_ocioso,
            'estado_biblioteca': 'Cerrada' if self.biblioteca_cerrada else 'Abierta',
            'cola': len(self.cola_espera),
            'ac_tiempo_permanencia': self.ac_tiempo_permanencia,
            'ac_clientes_leyendo': self.total_clientes_leyendo,
            'clientes': clientes_copiados
        }


# ==================== THREAD DE SIMULACIÓN ====================

class SimulacionThread(QThread):
    """Thread para ejecutar la simulación sin bloquear la UI"""
    finished = pyqtSignal(object)  # historial_filas
    error = pyqtSignal(str)
    progress = pyqtSignal(int)

    def __init__(self, simulacion: Simulacion):
        super().__init__()
        self.simulacion = simulacion

    def run(self):
        try:
            self.simulacion.iniciar() 
            
            while True:
                resultado = self.simulacion.ejecutar_paso()
                if resultado is None:
                    break

                # Usamos el número de fila como contador de eventos procesados (después de la inicialización)
                if self.simulacion.numero_fila % 10 == 0:
                    self.progress.emit(self.simulacion.numero_fila - 1) 

            self.finished.emit(self.simulacion.historial_filas)

        except Exception as e:
            self.error.emit(str(e))


# ==================== VENTANA DE CONFIGURACIÓN (MÍNIMA) ====================

class ConfiguracionWindow(QDialog):
    """Ventana para configurar los parámetros esenciales de la simulación."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("⚙️ Configuración de Parámetros Esenciales")
        self.setGeometry(100, 100, 700, 500)
        self.parametros = None

        main_layout = QVBoxLayout()
        self.setLayout(main_layout)

        tabs = QTabWidget()
        tabs.addTab(self.crear_tab_parametros(), "Parámetros Esenciales")
        tabs.addTab(self.crear_tab_fijos(), "Parámetros Fijos (Enunciado)")
        main_layout.addWidget(tabs)

        btn_iniciar = QPushButton("🚀 Iniciar Simulación con Parámetros")
        btn_iniciar.setStyleSheet("background-color: #2196F3; color: white; padding: 10px; font-size: 12pt; font-weight: bold; border-radius: 5px;")
        btn_iniciar.clicked.connect(self.aceptar)
        main_layout.addWidget(btn_iniciar)

    def crear_tab_parametros(self):
        widget = QWidget()
        layout = QFormLayout()
        widget.setLayout(layout)
        
        # 1. Frecuencia de Llegada (Default: 4 min)
        group_llegadas = QGroupBox("1. Frecuencia de Llegada")
        group_llegadas.setLayout(QFormLayout())
        self.s_tiempo_entre_llegadas = QDoubleSpinBox(value=4.0, decimals=1, minimum=0.1, maximum=30.0, singleStep=0.5)
        group_llegadas.layout().addRow("Tiempo Entre Llegadas (min):", self.s_tiempo_entre_llegadas)
        layout.addRow(group_llegadas)

        # 2. Probabilidades de Objetivo (Default: 45%, 45%, 10%)
        group_objetivo = QGroupBox("2. Probabilidades de Objetivo (Suma debe ser 1.0)")
        group_objetivo.setLayout(QFormLayout())
        self.s_prob_pedir = QDoubleSpinBox(value=0.45, decimals=2, minimum=0.0, maximum=1.0, singleStep=0.01)
        self.s_prob_devolver = QDoubleSpinBox(value=0.45, decimals=2, minimum=0.0, maximum=1.0, singleStep=0.01)
        self.s_prob_consultar = QDoubleSpinBox(value=0.10, decimals=2, minimum=0.0, maximum=1.0, singleStep=0.01)
        group_objetivo.layout().addRow("P(Pedir libro):", self.s_prob_pedir)
        group_objetivo.layout().addRow("P(Devolver libro):", self.s_prob_devolver)
        group_objetivo.layout().addRow("P(Consultar):", self.s_prob_consultar)
        layout.addRow(group_objetivo)
        
        # 3. Tiempos de Consulta (U[Min, Max])
        group_consulta = QGroupBox("3. Tiempo de Consulta (U[Min, Max])")
        group_consulta.setLayout(QFormLayout())
        self.s_cons_min = QDoubleSpinBox(value=2.0, decimals=1, minimum=0.1, maximum=10.0, singleStep=0.1)
        self.s_cons_max = QDoubleSpinBox(value=5.0, decimals=1, minimum=0.1, maximum=10.0, singleStep=0.1)
        group_consulta.layout().addRow("Consulta Min (min):", self.s_cons_min)
        group_consulta.layout().addRow("Consulta Max (min):", self.s_cons_max)
        layout.addRow(group_consulta)

        # 4. Decisión de Lectura (Default: 60% retira)
        group_lectura_decision = QGroupBox("4. Decisión de Retirarse/Leer")
        group_lectura_decision.setLayout(QFormLayout())
        self.s_prob_retirarse = QDoubleSpinBox(value=0.60, decimals=2, minimum=0.0, maximum=1.0, singleStep=0.01)
        group_lectura_decision.layout().addRow("P(Se retira después de pedir):", self.s_prob_retirarse)
        layout.addRow(group_lectura_decision)
        
        # 5. Constantes K para Integración Numérica
        group_K = QGroupBox("5. Constantes K para Integración (según rango de páginas)")
        group_K.setLayout(QFormLayout())
        self.s_K_100_200 = QSpinBox(value=100, minimum=1, maximum=500, singleStep=5)
        self.s_K_200_300 = QSpinBox(value=90, minimum=1, maximum=500, singleStep=5)
        self.s_K_300_plus = QSpinBox(value=70, minimum=1, maximum=500, singleStep=5)
        group_K.layout().addRow("K para [100-200] páginas:", self.s_K_100_200)
        group_K.layout().addRow("K para (200-300] páginas:", self.s_K_200_300)
        group_K.layout().addRow("K para (300+) páginas:", self.s_K_300_plus)
        layout.addRow(group_K)

        # 6. Número de Iteraciones Máximo
        group_iteraciones = QGroupBox("6. Iteraciones y Simulación")
        group_iteraciones.setLayout(QFormLayout())
        # Maximo de 100000 iteraciones
        self.s_max_iteraciones = QSpinBox(value=10000, minimum=1, maximum=100000, singleStep=1000)
        group_iteraciones.layout().addRow("Iteraciones Máx. (Eventos):", self.s_max_iteraciones)
        layout.addRow(group_iteraciones)

        return widget
    
    def crear_tab_fijos(self):
        widget = QWidget()
        layout = QFormLayout()
        widget.setLayout(layout)

        group_fijos = QGroupBox("Parámetros Fijos (No Parametrizables por Requerimiento)")
        group_fijos.setStyleSheet("QGroupBox { font-style: italic; } QLabel { color: #6c757d; }")

        fixed_layout = QFormLayout()
        fixed_layout.addRow("T. Búsqueda (Media EXP):", QLabel("6.0 min"))
        fixed_layout.addRow("T. Devolución (U[Min, Max]):", QLabel("U[1.5, 2.5] min"))
        fixed_layout.addRow("Páginas (U[Min, Max]):", QLabel("U[100, 350]"))
        fixed_layout.addRow("Capacidad Máxima:", QLabel("20 personas"))
        fixed_layout.addRow("T. Simulación Máximo:", QLabel("480.0 min (Solo para métricas de % tiempo cerrada)"))
        fixed_layout.addRow("Paso h (Euler):", QLabel("0.1 (Unidad = 10 min)"))

        group_fijos.setLayout(fixed_layout)
        layout.addRow(group_fijos)
        return widget


    def aceptar(self):
        """Valida y guarda los parámetros antes de cerrar."""
        p_pedir = self.s_prob_pedir.value()
        p_devolver = self.s_prob_devolver.value()
        p_consultar = self.s_prob_consultar.value()

        # Validación de la suma de probabilidades
        if not math.isclose(p_pedir + p_devolver + p_consultar, 1.0, abs_tol=0.001):
            QMessageBox.critical(self, "Error de Probabilidad", 
                                "La suma de las probabilidades de objetivo (Pedir, Devolver, Consultar) debe ser 1.0.")
            return
        
        # Validación de rangos
        if self.s_cons_min.value() > self.s_cons_max.value():
            QMessageBox.critical(self, "Error de Rango", 
                                "El valor mínimo de consulta no puede ser mayor que el máximo.")
            return

        self.parametros = {
            'tiempo_entre_llegadas': self.s_tiempo_entre_llegadas.value(),
            'prob_pedir_libro': p_pedir,
            'prob_devolver_libro': p_devolver,
            'prob_consultar': p_consultar,
            'tiempo_consulta_min': self.s_cons_min.value(),
            'tiempo_consulta_max': self.s_cons_max.value(),
            'prob_retirarse': self.s_prob_retirarse.value(),
            'K_100_200': self.s_K_100_200.value(),
            'K_200_300': self.s_K_200_300.value(),
            'K_300_plus': self.s_K_300_plus.value(),
            'max_iteraciones': self.s_max_iteraciones.value(),
        }
        self.accept()


# ==================== INTERFAZ GRÁFICA PRINCIPAL ====================

class MainWindow(QMainWindow):
    """Ventana principal de la aplicación"""
    
    # Límite de filas visibles
    MAX_FILAS_VISIBLES = 300 

    def __init__(self):
        super().__init__()
        self.simulacion: Optional[Simulacion] = None
        self.historial_filas = []
        self.thread = None
        self.parametros_simulacion = None 
        # Referencias a los QLabel para actualizar métricas
        self.lbl_metricas_dict: Dict[str, QLabel] = {}

        self.init_ui()
        # Inicia la ventana de configuración inmediatamente al arrancar
        self.mostrar_configuracion() 

    def crear_tab_metricas(self) -> QWidget:
        """Crea el widget de la pestaña de métricas de resumen."""
        widget = QWidget()
        main_layout = QVBoxLayout(widget)
        
        title = QLabel("📊 **RESUMEN DE MÉTRICAS DE SIMULACIÓN**")
        title.setFont(QFont("Arial", 14, QFont.Bold))
        title.setStyleSheet("padding: 10px; background-color: #f0f0f0; border-radius: 5px;")
        main_layout.addWidget(title)
        
        group = QGroupBox("Resultados Finales")
        group.setFont(QFont("Arial", 11, QFont.Bold))
        form_layout = QFormLayout()
        
        # Diccionario para almacenar los QLabel y poder actualizarlos
        metricas = {
            'eventos': ("Eventos Procesados:", "0"),
            'tiempo_simulado': ("Tiempo Simulado (min):", "0.00"),
            'clientes_generados': ("Clientes Generados:", "0"),
            'clientes_rechazados': ("Clientes Rechazados:", "0 (0.00%)"),
            'promedio_permanencia': ("Promedio de Permanencia (min):", "0.00"),
            'tiempo_cerrada': ("% Tiempo Cerrada (Base 480 min):", "0.00%"),
        }
        
        for key, (label_text, default_value) in metricas.items():
            lbl_value = QLabel(default_value)
            lbl_value.setFont(QFont("Arial", 10))
            lbl_value.setStyleSheet("color: #007bff; font-weight: bold;")
            form_layout.addRow(QLabel(label_text), lbl_value)
            self.lbl_metricas_dict[key] = lbl_value
        
        group.setLayout(form_layout)
        main_layout.addWidget(group)
        main_layout.addStretch(1) # Relleno para que no se pegue abajo
        
        return widget


    def init_ui(self):
        """Inicializa la interfaz"""
        self.setWindowTitle("Simulación de Biblioteca - Sistema de Eventos Discretos con Método de Euler")
        self.setGeometry(50, 50, 1800, 950)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout()
        central_widget.setLayout(main_layout)

        title = QLabel("🏛️ SIMULACIÓN DE BIBLIOTECA - EVENTOS DISCRETOS + INTEGRACIÓN DE EULER")
        title.setFont(QFont("Arial", 16, QFont.Bold))
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("background-color: #4472C4; color: white; padding: 15px; border-radius: 5px;")
        main_layout.addWidget(title)

        info_euler = QLabel("📐 Integración Numérica: dP/dt = K/5 (Método de Euler para calcular tiempo de lectura)")
        info_euler.setFont(QFont("Arial", 10))
        info_euler.setAlignment(Qt.AlignCenter)
        info_euler.setStyleSheet("background-color: #fff3cd; padding: 8px; border-radius: 3px; color: #856404;")
        main_layout.addWidget(info_euler)

        self.control_panel = self.crear_panel_control()
        main_layout.addWidget(self.control_panel)

        # CAMBIO: Usamos QTabWidget para la tabla y las métricas
        self.tab_widget = QTabWidget() 
        main_layout.addWidget(self.tab_widget)

        # 1. Pestaña de la Tabla (Eventos)
        self.tabla = QTableWidget()
        self.tabla.setAlternatingRowColors(True)
        self.tabla.setSelectionBehavior(QTableWidget.SelectRows)
        self.tabla.setSelectionMode(QTableWidget.SingleSelection)
        self.tabla.setStyleSheet("""
            QTableWidget { gridline-color: #d0d0d0; font-size: 9pt; background-color: white; }
            QHeaderView::section { background-color: #4472C4; color: white; font-weight: bold; padding: 6px; border: 1px solid #2a52a4; }
            QTableWidget::item:alternate { background-color: #f9f9f9; }
            QTableWidget::item:selected { background-color: #0078d7; color: white; }
        """)
        
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOn)
        scroll.setWidget(self.tabla)
        
        self.tab_widget.addTab(scroll, "📋 Tabla de Eventos")

        # 2. Pestaña de Métricas
        self.tab_metricas = self.crear_tab_metricas()
        self.tab_widget.addTab(self.tab_metricas, "📊 Métricas Finales")
        
        # El resto del layout se mantiene igual
        self.progress_bar = QProgressBar()
        self.progress_bar.setStyleSheet("QProgressBar { border: 2px solid #ccc; border-radius: 5px; text-align: center; } QProgressBar::chunk { background-color: #4CAF50; }")
        self.progress_bar.setVisible(False)
        main_layout.addWidget(self.progress_bar)

        status_layout = QHBoxLayout()
        self.lbl_status = QLabel("⏰ Esperando configuración...")
        self.lbl_status.setFont(QFont("Arial", 10, QFont.Bold))
        status_layout.addWidget(self.lbl_status)
        status_layout.addStretch()
        main_layout.addLayout(status_layout)
        
        self.set_botones_habilitados(False)

    def set_botones_habilitados(self, enabled: bool):
        self.btn_ejecutar.setEnabled(enabled)
        self.btn_exportar.setEnabled(False)
        self.btn_reiniciar.setEnabled(True)

    def crear_panel_control(self) -> QGroupBox:
        """Crea el panel de control"""
        group = QGroupBox("🎮 Controles de Simulación")
        group.setFont(QFont("Arial", 11, QFont.Bold))
        layout = QHBoxLayout()

        self.btn_ejecutar = QPushButton("▶ Ejecutar Simulación Completa")
        self.btn_ejecutar.setStyleSheet("background-color: #4CAF50; color: white; padding: 12px 24px; font-size: 13pt; font-weight: bold; border-radius: 5px;")
        self.btn_ejecutar.clicked.connect(self.ejecutar_simulacion)
        layout.addWidget(self.btn_ejecutar)

        self.btn_exportar = QPushButton("📊 Exportar a Excel")
        self.btn_exportar.setStyleSheet("background-color: #2196F3; color: white; padding: 12px 24px; font-size: 13pt; font-weight: bold; border-radius: 5px;")
        self.btn_exportar.clicked.connect(self.exportar_excel)
        layout.addWidget(self.btn_exportar)

        self.btn_reiniciar = QPushButton("🔄 Reconfigurar / Nueva Simulación")
        self.btn_reiniciar.setStyleSheet("background-color: #F44336; color: white; padding: 12px 24px; font-size: 13pt; font-weight: bold; border-radius: 5px;")
        self.btn_reiniciar.clicked.connect(self.mostrar_configuracion)
        layout.addWidget(self.btn_reiniciar)

        layout.addStretch()
        group.setLayout(layout)
        return group
    
    def mostrar_configuracion(self):
        """Abre la ventana de diálogo para configurar parámetros."""
        
        # Esta llamada necesita que ConfiguracionWindow esté definida arriba
        config_dialog = ConfiguracionWindow(self) 
        
        if config_dialog.exec_() == QDialog.Accepted:
            self.parametros_simulacion = config_dialog.parametros
            self.lbl_status.setText(f"✅ Configuración cargada. Llegan cada {self.parametros_simulacion['tiempo_entre_llegadas']:.1f} min. Iteraciones Máx: {self.parametros_simulacion['max_iteraciones']}")
            self.set_botones_habilitados(True)
            self.reiniciar_tabla()
        else:
            self.lbl_status.setText("❌ Configuración cancelada. Presione 'Reconfigurar / Nueva Simulación'.")
            self.set_botones_habilitados(False)


    def ejecutar_simulacion(self):
        """Ejecuta la simulación completa"""
        if not self.parametros_simulacion:
            QMessageBox.warning(self, "Advertencia", "Debe configurar y cargar los parámetros primero.")
            self.mostrar_configuracion()
            return
            
        self.reiniciar_tabla()

        self.btn_ejecutar.setEnabled(False)
        self.btn_exportar.setEnabled(False)
        self.progress_bar.setVisible(True)
        
        max_iter = self.parametros_simulacion.get('max_iteraciones', 10000)
        self.progress_bar.setRange(0, max_iter)
        
        self.lbl_status.setText("⏳ Ejecutando simulación completa...")

        # Inyectamos los parámetros en la simulación
        self.simulacion = Simulacion(self.parametros_simulacion)

        self.thread = SimulacionThread(self.simulacion)
        self.thread.finished.connect(self.simulacion_completada)
        self.thread.error.connect(self.simulacion_error)
        self.thread.progress.connect(self.actualizar_progreso)
        self.thread.start()

    def actualizar_progreso(self, eventos):
        """Actualiza el progreso"""
        max_iter = self.parametros_simulacion.get('max_iteraciones', 10000)
        self.progress_bar.setValue(eventos)
        self.lbl_status.setText(f"⏳ Procesando... Evento {eventos} de {max_iter} (sin contar Inicialización)")

    def actualizar_metricas_ui(self, total_eventos_simulados: int, total_clientes: int, promedio_permanencia: float, porcentaje_rechazados: float, porcentaje_tiempo_cerrada: float):
        """Actualiza los QLabel en la pestaña de Métricas."""
        
        reloj_final = self.simulacion.reloj if self.simulacion else 0.0
        rechazados = self.simulacion.total_clientes_rechazados if self.simulacion else 0
        
        self.lbl_metricas_dict['eventos'].setText(str(total_eventos_simulados))
        self.lbl_metricas_dict['tiempo_simulado'].setText(f"{reloj_final:.2f} min")
        self.lbl_metricas_dict['clientes_generados'].setText(str(total_clientes))
        self.lbl_metricas_dict['clientes_rechazados'].setText(f"{rechazados} ({porcentaje_rechazados:.2f}%)")
        self.lbl_metricas_dict['promedio_permanencia'].setText(f"{promedio_permanencia:.2f} min")
        self.lbl_metricas_dict['tiempo_cerrada'].setText(f"{porcentaje_tiempo_cerrada:.2f}%")
        
        # Seleccionar la pestaña de métricas para mostrar el resultado
        self.tab_widget.setCurrentIndex(1)

    def simulacion_completada(self, historial_filas):
        """Callback cuando termina la simulación"""
        self.historial_filas = historial_filas
        self.progress_bar.setVisible(False)
        self.set_botones_habilitados(True)
        self.btn_ejecutar.setEnabled(True)
        self.btn_exportar.setEnabled(True)

        self.poblar_tabla()
        
        # --- CÁLCULO DE MÉTRICAS ---
        
        # Restamos 1 por la fila de Inicialización y 1 por la fila de Fin Simulación forzado.
        total_eventos_simulados = len(historial_filas) - 2
        if total_eventos_simulados < 0:
            total_eventos_simulados = 0 
        
        # CÁLCULO DE MÉTRICAS FINALES
        total_clientes = self.simulacion.total_clientes_generados
        promedio_permanencia = (self.simulacion.ac_tiempo_permanencia / 
                                self.simulacion.total_clientes_atendidos) if self.simulacion.total_clientes_atendidos > 0 else 0
        
        porcentaje_rechazados = (self.simulacion.total_clientes_rechazados / 
                                 total_clientes) * 100 if total_clientes > 0 else 0
        
        porcentaje_tiempo_cerrada = (self.simulacion.tiempo_biblioteca_cerrada_ac / 
                                      self.simulacion.tiempo_maximo) * 100 if self.simulacion.tiempo_maximo > 0 else 0

        # Actualizamos la pestaña de Métricas
        self.actualizar_metricas_ui(
            total_eventos_simulados, 
            total_clientes, 
            promedio_permanencia, 
            porcentaje_rechazados, 
            porcentaje_tiempo_cerrada
        )
        
        # Actualizamos la barra de estado inferior
        self.lbl_status.setText(f"✅ Simulación completa: {total_eventos_simulados} eventos | "
                                f"T. Sim: {self.simulacion.reloj:.2f} min | "
                                f"% Rechazados: {porcentaje_rechazados:.2f}%")

    def simulacion_error(self, error_msg):
        """Callback cuando hay error"""
        self.progress_bar.setVisible(False)
        self.set_botones_habilitados(True)
        self.btn_ejecutar.setEnabled(True)
        self.lbl_status.setText("❌ Error en la simulación")
        QMessageBox.critical(self, "Error", f"Error en la simulación:\n\n{error_msg}")

    def reiniciar_tabla(self):
        """Limpia los datos y la tabla de la simulación anterior."""
        self.simulacion = None
        self.historial_filas = []
        self.tabla.setRowCount(0)
        self.tabla.setColumnCount(0)
        self.btn_exportar.setEnabled(False)
        self.lbl_status.setText("⏰ Esperando...")


    def poblar_tabla(self):
        if not self.historial_filas: return

        todos_clientes_ids = set()
        # Se sigue iterando sobre todas las filas para obtener todos los IDs de clientes
        for fila in self.historial_filas[1:]: 
            for cliente in fila['clientes']:
                if cliente.estado != "RECHAZADO":
                    todos_clientes_ids.add(cliente.id)
        clientes_ordenados = sorted(list(todos_clientes_ids))

        columnas_fijas = [
            "n°", "Evento", "Reloj",
            "T.entre llegadas", "Próxima llegada", 
            "RND obj", "Objetivo", 
            "RND búsq", "T.búsqueda (EXP)", "fin_búsq_emp1", "fin_búsq_emp2",
            "RND decisión", "Se retira?", "RND pág", "Páginas", "T.lectura (Euler)",
            "RND dev", "T.devolución", "fin_dev_emp1", "fin_dev_emp2",
            "RND cons", "T.consulta", "fin_cons",
            "Emp1", "AC at1", "AC oc1", "Emp2", "AC at2", "AC oc2",
            "Cola", "AC perm", "AC leyendo"
        ]

        columnas_clientes = []
        for cid in clientes_ordenados:
            columnas_clientes.extend([
                f"C{cid} Estado", f"C{cid} Objetivo", f"C{cid} Hora", f"C{cid} T.at", f"C{cid} Fin lect"
            ])

        todas_columnas = columnas_fijas + columnas_clientes

        self.tabla.setColumnCount(len(todas_columnas))
        self.tabla.setHorizontalHeaderLabels(todas_columnas)
        
        # Determina cuántas filas se van a mostrar en la interfaz (máximo 300)
        filas_a_mostrar = min(len(self.historial_filas), self.MAX_FILAS_VISIBLES)
        self.tabla.setRowCount(filas_a_mostrar) # Establece solo el número máximo visible

        header = self.tabla.horizontalHeader()
        header.setDefaultSectionSize(85)
        header.setSectionResizeMode(QHeaderView.Interactive)

        # Solo itera sobre las filas que se van a mostrar
        for row in range(filas_a_mostrar):
            fila = self.historial_filas[row]
            self.agregar_fila(row, fila, clientes_ordenados)

        # Muestra un mensaje si se truncó la tabla visible
        if len(self.historial_filas) > self.MAX_FILAS_VISIBLES:
            self.lbl_status.setText(
                f"⚠️ Simulación completa: {len(self.historial_filas)} eventos. "
                f"Solo mostrando las primeras {self.MAX_FILAS_VISIBLES} filas. ¡Exporta a Excel para ver todas!"
            )


    def agregar_fila(self, row: int, datos: dict, clientes_ordenados: List[int]):
        def fmt(val):
            if val == '' or val is None: return ''
            if isinstance(val, float): return f"{val:.2f}"
            return str(val)

        reloj_actual = datos['reloj']
        clientes_dict = {c.id: c for c in datos['clientes']}

        # Recolectar todos los tiempos futuros para identificar el próximo evento
        tiempos_futuros = []

        # Próxima llegada
        prox_llegada_val = datos.get('proxima_llegada')
        if isinstance(prox_llegada_val, (int, float)) and prox_llegada_val > reloj_actual:
            tiempos_futuros.append(prox_llegada_val)

        # Fin de atenciones
        for key in ['fin_atencion_alq1', 'fin_atencion_alq2', 'fin_atencion_dev1', 'fin_atencion_dev2', 'fin_atencion_cons']:
            t_fin = datos.get(key)
            if isinstance(t_fin, (int, float)) and t_fin > reloj_actual:
                tiempos_futuros.append(t_fin)

        # Fin de lecturas
        for cliente in datos['clientes']:
            if cliente.fin_lectura is not None and cliente.fin_lectura > reloj_actual:
                tiempos_futuros.append(cliente.fin_lectura)

        # Identificar el mínimo tiempo futuro (próximo evento)
        min_tiempo_futuro = min(tiempos_futuros) if tiempos_futuros else None

        COL_PROX_LLEGADA = 4
        COL_FIN_BUSQ_EMP1 = 9
        COL_FIN_BUSQ_EMP2 = 10
        COL_FIN_DEV_EMP1 = 18
        COL_FIN_DEV_EMP2 = 19
        COL_FIN_CONS = 22

        es_llegada = datos['evento'].startswith(TipoEvento.LLEGADA_CLIENTE.value)
        es_rechazado = datos['objetivo'] == 'RECHAZADO'
        mostrar_rnd_obj = es_llegada and not es_rechazado

        valores = [
            datos['n'], datos['evento'], fmt(datos['reloj']), fmt(datos['tiempo_entre_llegadas']), fmt(datos['proxima_llegada']),
            fmt(datos['rnd_objetivo']) if mostrar_rnd_obj else '', datos['objetivo'] if es_llegada else '',
            fmt(datos['rnd_busqueda']), fmt(datos['tiempo_busqueda']), fmt(datos['fin_atencion_alq1']), fmt(datos['fin_atencion_alq2']),
            fmt(datos['rnd_decision']), datos['se_retira'], fmt(datos['rnd_paginas']), datos['paginas'], fmt(datos['tiempo_lectura']),
            fmt(datos['rnd_devolucion']), fmt(datos['tiempo_devolucion']), fmt(datos['fin_atencion_dev1']), fmt(datos['fin_atencion_dev2']),
            fmt(datos['rnd_consulta']), fmt(datos['tiempo_consulta']), fmt(datos['fin_atencion_cons']),
            datos['empleado1_estado'], fmt(datos['empleado1_ac_atencion']), fmt(datos['empleado1_ac_ocioso']),
            datos['empleado2_estado'], fmt(datos['empleado2_ac_atencion']), fmt(datos['empleado2_ac_ocioso']),
            str(datos['cola']), fmt(datos['ac_tiempo_permanencia']), str(datos['ac_clientes_leyendo'])
        ]

        # Mapa de tiempos programados en cada columna
        # Estos tiempos se resaltan en rojo cuando coinciden con el reloj (evento que acaba de ocurrir)
        map_tiempos_eventos = {
            COL_PROX_LLEGADA: datos.get('proxima_llegada'),
            COL_FIN_BUSQ_EMP1: datos.get('fin_atencion_alq1'),
            COL_FIN_BUSQ_EMP2: datos.get('fin_atencion_alq2'),
            COL_FIN_DEV_EMP1: datos.get('fin_atencion_dev1'),
            COL_FIN_DEV_EMP2: datos.get('fin_atencion_dev2'),
            COL_FIN_CONS: datos.get('fin_atencion_cons'),
        }
        
        col_offset = len(valores)

        def obtener_objetivo_cliente(cliente) -> str:
            """Retorna una descripción del objetivo/acción del cliente"""
            if cliente.estado == "RECHAZADO":
                return "RECHAZADO"
            elif cliente.estado == "En cola":
                # Distinguir si está esperando por primera vez o esperando para devolver
                if cliente.objetivo == TipoObjetivo.DEVOLVER_LIBRO:
                    return "Esperando devolver"
                else:
                    return "En espera"
            elif cliente.estado == "Siendo atendido":
                if cliente.objetivo == TipoObjetivo.PEDIR_LIBRO:
                    return "Pidiendo libro"
                elif cliente.objetivo == TipoObjetivo.DEVOLVER_LIBRO:
                    return "Devolviendo"
                elif cliente.objetivo == TipoObjetivo.CONSULTAR:
                    return "Consultando"
            elif cliente.estado == "Leyendo":
                return "Leyendo"
            elif cliente.estado == "Fuera del sistema":
                if cliente.se_retira:
                    return "Se retiró"
                else:
                    return "Finalizó"
            else:
                return cliente.objetivo.value if cliente.objetivo else ""

        for cid in clientes_ordenados:
            if cid in clientes_dict:
                c = clientes_dict[cid]
                tiempo_at = next((t for t in [c.tiempo_busqueda, c.tiempo_devolucion, c.tiempo_consulta] if t > 0), '')
                objetivo_desc = obtener_objetivo_cliente(c)
                valores.extend([c.estado, objetivo_desc, fmt(c.hora_llegada), fmt(tiempo_at) if tiempo_at else '', fmt(c.fin_lectura) if c.fin_lectura else ''])
                # Agregar columna de fin_lectura al mapa para resaltado
                if c.fin_lectura is not None:
                    map_tiempos_eventos[col_offset + 4] = c.fin_lectura  # Ajuste por la nueva columna
            else:
                valores.extend(['', '', '', '', ''])
            col_offset += 5  # Ahora son 5 columnas por cliente

        # Tolerancia para comparaciones de punto flotante
        EPSILON = 0.001

        for col, valor in enumerate(valores):
            item = QTableWidgetItem(str(valor))
            item.setTextAlignment(Qt.AlignCenter)

            color_fondo = QColor(249, 249, 249) if row % 2 == 0 else QColor(255, 255, 255)

            if col == 1:
                if 'inicializacion' in str(valor).lower(): color_fondo = QColor(230, 230, 250)
                elif 'llegada' in str(valor).lower(): color_fondo = QColor(200, 230, 201)
                elif 'fin_atencion' in str(valor).lower(): color_fondo = QColor(255, 224, 178)
                elif 'fin_lectura' in str(valor).lower(): color_fondo = QColor(187, 222, 251)

            if es_llegada and es_rechazado and col == 6:
                color_fondo = QColor(255, 199, 206)

            item.setBackground(color_fondo)

            # RESALTAR EN ROJO: El próximo evento a ocurrir (mínimo tiempo futuro > reloj)
            tiempo_columna = map_tiempos_eventos.get(col)
            if tiempo_columna is not None and isinstance(tiempo_columna, (int, float)):
                # Si este tiempo es el próximo evento (mínimo tiempo futuro)
                if min_tiempo_futuro is not None and abs(tiempo_columna - min_tiempo_futuro) < EPSILON:
                    item.setForeground(QColor(255, 0, 0))
                    item.setFont(QFont("Arial", 9, QFont.Bold))

            self.tabla.setItem(row, col, item)


    def exportar_excel(self):
        
        if not OPENPYXL_DISPONIBLE:
            QMessageBox.warning(self, "Advertencia", "La librería 'openpyxl' no está instalada.\nInstálala con: pip install openpyxl")
            return
        if not self.historial_filas:
            QMessageBox.warning(self, "Advertencia", "No hay datos para exportar")
            return

        filename, _ = QFileDialog.getSaveFileName(self, "Guardar archivo Excel", "simulacion_biblioteca.xlsx", "Excel Files (*.xlsx)")
        if not filename: return

        try:
            wb = openpyxl.Workbook()
            
            # 1. Preparar datos y encabezados para la exportación completa
            todos_clientes_ids = set()
            for fila in self.historial_filas[1:]: 
                for cliente in fila['clientes']:
                    if cliente.estado != "RECHAZADO":
                        todos_clientes_ids.add(cliente.id)
            clientes_ordenados = sorted(list(todos_clientes_ids))
            
            # Tomar encabezados de la tabla
            if self.tabla.columnCount() == 0:
                QMessageBox.warning(self, "Advertencia", "La tabla no se ha inicializado correctamente. Ejecute la simulación primero.")
                return

            headers = [self.tabla.horizontalHeaderItem(col).text() for col in range(self.tabla.columnCount())]
            
            # --- HOJA 1: SIMULACIÓN DE EVENTOS (TODAS LAS FILAS) ---
            ws = wb.active
            ws.title = "Simulación"
            
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = ExcelFont(bold=True, color="FFFFFF")
            center_align = Alignment(horizontal="center", vertical="center")

            # Escribir encabezados
            for col_idx, header in enumerate(headers):
                cell = ws.cell(row=1, column=col_idx + 1)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align

            # Escribir TODAS las filas de historial_filas
            for row_idx, fila_data in enumerate(self.historial_filas):
                
                # RECREAR LOS VALORES DE LA FILA COMPLETA
                
                def fmt_excel(val):
                    if val == '' or val is None: return ''
                    if isinstance(val, float): return f"{val:.2f}"
                    return str(val)

                # Identificar el cliente actual (para rellenar los RND/Tiempos)
                cliente_id_evento = fila_data['evento'].split(' C')[-1].split(' F')[-1].strip()
                try:
                    cliente_id_evento = int(cliente_id_evento)
                except ValueError:
                    cliente_id_evento = None

                cliente_actual = None
                if cliente_id_evento is not None:
                    cliente_actual = next((c for c in fila_data['clientes'] if c.id == cliente_id_evento), None)
                
                if cliente_actual and cliente_actual.estado == 'RECHAZADO':
                    cliente_actual = None 
                
                # Generar valores para columnas fijas
                rnd_busqueda = cliente_actual.rnd_tiempo_busqueda if cliente_actual and cliente_actual.objetivo == TipoObjetivo.PEDIR_LIBRO and cliente_actual.rnd_tiempo_busqueda > 0 else ''
                tiempo_busqueda = cliente_actual.tiempo_busqueda if cliente_actual and cliente_actual.objetivo == TipoObjetivo.PEDIR_LIBRO and cliente_actual.tiempo_busqueda > 0 else ''
                rnd_devolucion = cliente_actual.rnd_tiempo_devolucion if cliente_actual and cliente_actual.objetivo == TipoObjetivo.DEVOLVER_LIBRO and cliente_actual.rnd_tiempo_devolucion > 0 else ''
                tiempo_devolucion = cliente_actual.tiempo_devolucion if cliente_actual and cliente_actual.objetivo == TipoObjetivo.DEVOLVER_LIBRO and cliente_actual.rnd_tiempo_devolucion > 0 else ''
                rnd_consulta = cliente_actual.rnd_tiempo_consulta if cliente_actual and cliente_actual.objetivo == TipoObjetivo.CONSULTAR and cliente_actual.rnd_tiempo_consulta > 0 else ''
                tiempo_consulta = cliente_actual.tiempo_consulta if cliente_actual and cliente_actual.objetivo == TipoObjetivo.CONSULTAR and cliente_actual.tiempo_consulta > 0 else ''

                es_llegada = fila_data['evento'].startswith("llegada_cliente")
                es_rechazado = fila_data.get('objetivo') == 'RECHAZADO'
                mostrar_rnd_obj = es_llegada and not es_rechazado
                
                valores_fijos = [
                    fila_data['n'], fila_data['evento'], fmt_excel(fila_data['reloj']), fmt_excel(fila_data['tiempo_entre_llegadas']), fmt_excel(fila_data['proxima_llegada']),
                    fmt_excel(fila_data['rnd_objetivo']) if mostrar_rnd_obj else '', fila_data['objetivo'] if es_llegada else '',       
                    fmt_excel(rnd_busqueda), fmt_excel(tiempo_busqueda), fmt_excel(fila_data['fin_atencion_alq1']), fmt_excel(fila_data['fin_atencion_alq2']),
                    fmt_excel(fila_data['rnd_decision']), fila_data['se_retira'], fmt_excel(fila_data['rnd_paginas']), fila_data['paginas'], fmt_excel(fila_data['tiempo_lectura']),
                    fmt_excel(rnd_devolucion), fmt_excel(tiempo_devolucion), fmt_excel(fila_data['fin_atencion_dev1']), fmt_excel(fila_data['fin_atencion_dev2']),
                    fmt_excel(rnd_consulta), fmt_excel(tiempo_consulta), fmt_excel(fila_data['fin_atencion_cons']),
                    fila_data['empleado1_estado'], fmt_excel(fila_data['empleado1_ac_atencion']), fmt_excel(fila_data['empleado1_ac_ocioso']), 
                    fila_data['empleado2_estado'], fmt_excel(fila_data['empleado2_ac_atencion']), fmt_excel(fila_data['empleado2_ac_ocioso']), 
                    str(fila_data['cola']), fmt_excel(fila_data['ac_tiempo_permanencia']), str(fila_data['ac_clientes_leyendo']) 
                ]
                
                # Función auxiliar para obtener objetivo del cliente (igual que en agregar_fila)
                def obtener_objetivo_cliente_excel(cliente) -> str:
                    if cliente.estado == "RECHAZADO":
                        return "RECHAZADO"
                    elif cliente.estado == "En cola":
                        # Distinguir si está esperando por primera vez o esperando para devolver
                        if cliente.objetivo == TipoObjetivo.DEVOLVER_LIBRO:
                            return "Esperando devolver"
                        else:
                            return "En espera"
                    elif cliente.estado == "Siendo atendido":
                        if cliente.objetivo == TipoObjetivo.PEDIR_LIBRO:
                            return "Pidiendo libro"
                        elif cliente.objetivo == TipoObjetivo.DEVOLVER_LIBRO:
                            return "Devolviendo"
                        elif cliente.objetivo == TipoObjetivo.CONSULTAR:
                            return "Consultando"
                    elif cliente.estado == "Leyendo":
                        return "Leyendo"
                    elif cliente.estado == "Fuera del sistema":
                        if cliente.se_retira:
                            return "Se retiró"
                        else:
                            return "Finalizó"
                    else:
                        return cliente.objetivo.value if cliente.objetivo else ""

                # Generar valores para columnas dinámicas de clientes
                clientes_dict = {c.id: c for c in fila_data['clientes']}
                valores_clientes = []
                for cid in clientes_ordenados:
                    if cid in clientes_dict:
                        c = clientes_dict[cid]
                        tiempo_at = next((t for t in [c.tiempo_busqueda, c.tiempo_devolucion, c.tiempo_consulta] if t > 0), '')
                        objetivo_desc = obtener_objetivo_cliente_excel(c)
                        valores_clientes.extend([c.estado, objetivo_desc, fmt_excel(c.hora_llegada), fmt_excel(tiempo_at) if tiempo_at else '', fmt_excel(c.fin_lectura) if c.fin_lectura else ''])
                    else:
                        valores_clientes.extend(['', '', '', '', ''])
                
                fila_completa = valores_fijos + valores_clientes
                
                for col_idx, valor in enumerate(fila_completa):
                    cell = ws.cell(row=row_idx + 2, column=col_idx + 1)
                    
                    try:
                        # Intenta convertir a número si es posible
                        text = str(valor).replace(',', '.')
                        if text and text.replace('.', '', 1).isdigit():
                            cell.value = float(text)
                            cell.number_format = '0.00' if '.' in text else '0'
                        else:
                            cell.value = valor
                    except ValueError:
                        cell.value = valor

                    cell.alignment = center_align

            # --- HOJA 2: TABLAS DE EULER ---
            if self.simulacion and self.simulacion.tablas_euler_clientes:
                ws_euler = wb.create_sheet("Tablas Euler")
                row_offset = 1
                
                euler_header_fill = PatternFill(start_color="8FAADC", end_color="8FAADC", fill_type="solid")

                for cliente_id, tabla in self.simulacion.tablas_euler_clientes.items():
                    if not tabla: continue

                    # Buscar páginas del cliente (solo para mostrar como referencia)
                    paginas_cliente = 0
                    for fila in self.historial_filas:
                        clientes = fila.get('clientes', [])
                        cliente_data = next((c for c in clientes if c.id == cliente_id), None)
                        if cliente_data and cliente_data.paginas_a_leer > 0:
                            paginas_cliente = cliente_data.paginas_a_leer
                            break
                    
                    # Título del cliente y parámetros
                    ws_euler.cell(row=row_offset, column=1, value=f"Cliente C{cliente_id}").font = ExcelFont(bold=True)
                    ws_euler.cell(row=row_offset, column=3, value=f"Páginas Objetivo: {paginas_cliente}").font = ExcelFont(italic=True)
                    ws_euler.cell(row=row_offset, column=5, value=f"K={self.simulacion.determinar_K(paginas_cliente)}").font = ExcelFont(italic=True)
                    ws_euler.cell(row=row_offset, column=7, value=f"h={self.simulacion.h_euler:.2f}").font = ExcelFont(italic=True)
                    row_offset += 1
                    
                    # Encabezados de la tabla
                    headers_euler = ['t (Tiempo)', 'p (Páginas Acumuladas)', 'dp/dt (Tasa de Lectura)']
                    for col_idx, header_euler in enumerate(headers_euler):
                        cell = ws_euler.cell(row=row_offset, column=col_idx + 1, value=header_euler)
                        cell.fill = euler_header_fill
                        cell.font = ExcelFont(bold=True)
                        cell.alignment = center_align
                    row_offset += 1
                    
                    # Datos de la tabla
                    for row_data in tabla:
                        for col_idx, key in enumerate(['t', 'p', 'dp/dt']):
                            val = row_data.get(key)
                            cell = ws_euler.cell(row=row_offset, column=col_idx + 1, value=val)
                            cell.number_format = '0.0000'
                            cell.alignment = Alignment(horizontal="right")
                        row_offset += 1
                    
                    row_offset += 2 # Espacio entre tablas
                
                # Ajustar ancho de columnas para la tabla Euler
                for col in ws_euler.columns:
                    max_length = 0
                    column = col[0].column_letter # Obtener la letra de la columna.
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws_euler.column_dimensions[column].width = adjusted_width

            wb.save(filename)
            QMessageBox.information(self, "Éxito", f"✅ Archivo exportado correctamente:\n{filename}")

        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al exportar:\n\n{str(e)}")

    def reiniciar(self):
        self.mostrar_configuracion()


def main():
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()